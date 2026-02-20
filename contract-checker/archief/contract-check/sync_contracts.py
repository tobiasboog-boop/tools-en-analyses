#!/usr/bin/env python3
"""
Sync contracts metadata from CSV or XLSX to database.

Usage:
    python sync_contracts.py contracts.xlsx
    python sync_contracts.py contracts.csv

Expected columns:
    Filename*, Client_ID*, Client_Name, Contract_Number,
    Start_Date, End_Date, Contract_Type, Notes

The 'filename' column is the unique key (must match SharePoint filename).
Existing records are updated, new records are inserted, missing records are soft-deleted.
"""
import argparse
import csv
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

from src.models import db, Contract, ContractChange


def parse_date(value):
    """Parse date value to date object."""
    if value is None:
        return None
    # Already a datetime/date object (from xlsx)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # String value (from csv)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def normalize_column(name: str) -> str:
    """Normalize column name: lowercase, strip, remove asterisk."""
    return name.lower().strip().rstrip("*").strip()


def load_xlsx(filepath: str) -> list[dict]:
    """Load contracts from XLSX file."""
    records = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Use first sheet
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        wb.close()
        return records

    # First row is header
    headers = [normalize_column(str(h)) if h else "" for h in rows[0]]

    for row in rows[1:]:
        if not any(row):  # Skip empty rows
            continue
        record = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                # Convert value to string if needed (except dates)
                if value is not None and not isinstance(value, (datetime, date)):
                    value = str(value).strip()
                record[headers[i]] = value
        records.append(record)

    wb.close()
    return records


def load_csv(filepath: str) -> list[dict]:
    """Load contracts from CSV file."""
    records = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            # Normalize column names
            normalized = {
                normalize_column(k): v.strip() if v else ""
                for k, v in row.items()
            }
            records.append(normalized)
    return records


def load_file(filepath: str) -> list[dict]:
    """Load contracts from CSV or XLSX file."""
    ext = Path(filepath).suffix.lower()
    if ext == ".xlsx":
        return load_xlsx(filepath)
    elif ext == ".csv":
        return load_csv(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .csv or .xlsx")


# Fields to track for change detection
TRACKED_FIELDS = [
    "client_id", "client_name", "contract_number",
    "start_date", "end_date", "contract_type", "notes"
]


def detect_changes(old_contract: Contract, new_data: dict) -> list[str]:
    """Detect which fields changed between old and new values."""
    changed = []
    for field in TRACKED_FIELDS:
        old_val = getattr(old_contract, field, None)
        new_val = new_data.get(field)
        # Normalize empty strings to None for comparison
        if old_val == "":
            old_val = None
        if new_val == "":
            new_val = None
        if old_val != new_val:
            changed.append(field)
    return changed


def log_change(session, contract_id: int, filename: str, change_type: str,
               old_client_id: str = None, new_client_id: str = None,
               changed_fields: list = None):
    """Log a contract change to the audit table."""
    change = ContractChange(
        contract_id=contract_id,
        filename=filename,
        change_type=change_type,
        old_client_id=old_client_id,
        new_client_id=new_client_id,
        changed_fields=changed_fields,
        changed_at=datetime.utcnow(),
        changed_by="sync_contracts"
    )
    session.add(change)


def sync_contracts(filepath: str, dry_run: bool = False):
    """
    Sync contracts from CSV/XLSX to database.

    Logic:
    - (filename, client_id) is the composite unique key
    - Same contract file can cover multiple clients (group contracts)
    - If (filename, client_id) exists in DB: update the record
    - If (filename, client_id) not in DB: insert new record
    - If (filename, client_id) in DB but not in file: soft-delete
    """
    print(f"Loading contracts from: {filepath}")
    csv_records = load_file(filepath)
    print(f"Found {len(csv_records)} records")

    if not csv_records:
        print("No records found. Exiting.")
        return

    # Validate required columns
    required = ["filename", "client_id"]
    first_row = csv_records[0]
    missing = [col for col in required if col not in first_row]
    if missing:
        print(f"ERROR: Missing required columns: {missing}")
        print(f"Available columns: {list(first_row.keys())}")
        sys.exit(1)

    session = db()
    try:
        # Get all existing contracts keyed by (filename, client_id)
        existing = {
            (c.filename, c.client_id): c
            for c in session.query(Contract).all()
        }
        # Track which (filename, client_id) pairs are in the import file
        csv_keys = {
            (r["filename"], r.get("client_id", ""))
            for r in csv_records
        }

        stats = {"inserted": 0, "updated": 0, "deleted": 0, "reactivated": 0}

        for row in csv_records:
            filename = row["filename"]
            client_id = row.get("client_id", "")
            if not filename or not client_id:
                print(f"  SKIP: empty filename or client_id")
                continue

            key = (filename, client_id)

            contract_data = {
                "filename": filename,
                "client_id": client_id,
                "client_name": row.get("client_name"),
                "contract_number": row.get("contract_number"),
                "start_date": parse_date(row.get("start_date", "")),
                "end_date": parse_date(row.get("end_date", "")),
                "contract_type": row.get("contract_type"),
                "notes": row.get("notes"),
                "last_synced_at": datetime.utcnow(),
            }

            if key in existing:
                # Update existing record
                contract = existing[key]
                was_inactive = not contract.active
                old_client_id = contract.client_id

                # Detect what changed (excluding client_id since it's part of key)
                changed_fields = detect_changes(contract, contract_data)

                for k, value in contract_data.items():
                    setattr(contract, k, value)

                contract.active = True
                contract.deleted_at = None
                contract.updated_at = datetime.utcnow()

                if was_inactive:
                    stats["reactivated"] += 1
                    print(f"  REACTIVATE: {filename} (client: {client_id})")
                    log_change(
                        session, contract.id, filename, "REACTIVATE",
                        old_client_id, client_id, changed_fields
                    )
                elif changed_fields:
                    stats["updated"] += 1
                    print(f"  UPDATE: {filename} (client: {client_id})")
                    print(f"    Changed: {', '.join(changed_fields)}")
                    log_change(
                        session, contract.id, filename, "UPDATE",
                        old_client_id, client_id, changed_fields
                    )
                # else: no changes, skip logging
            else:
                # Insert new record
                contract = Contract(**contract_data)
                session.add(contract)
                session.flush()  # Get the ID for logging
                stats["inserted"] += 1
                print(f"  INSERT: {filename} (client: {client_id})")
                log_change(
                    session, contract.id, filename, "INSERT",
                    None, client_id, None
                )

        # Soft-delete contracts not in import file
        for key, contract in existing.items():
            if key not in csv_keys and contract.active:
                contract.active = False
                contract.deleted_at = datetime.utcnow()
                stats["deleted"] += 1
                print(f"  DEACTIVATE: {contract.filename} (client: {contract.client_id})")
                log_change(
                    session, contract.id, contract.filename, "DEACTIVATE",
                    contract.client_id, None, None
                )

        if dry_run:
            print("\n[DRY RUN] Rolling back changes...")
            session.rollback()
        else:
            session.commit()
            print("\nChanges committed to database.")

        print("\nSummary:")
        print(f"  Inserted:    {stats['inserted']}")
        print(f"  Updated:     {stats['updated']}")
        print(f"  Reactivated: {stats['reactivated']}")
        print(f"  Deleted:     {stats['deleted']}")

    except Exception as e:
        session.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        session.close()


def main():
    parser = argparse.ArgumentParser(
        description="Sync contracts metadata from CSV/XLSX to database"
    )
    parser.add_argument(
        "file",
        help="Path to CSV or XLSX file with contract metadata"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would be changed without committing"
    )

    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"ERROR: File not found: {args.file}")
        sys.exit(1)

    sync_contracts(args.file, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
